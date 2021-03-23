# -*- coding:utf-8 -*- 
import os
import sys
import datetime
import math
from openpyxl import Workbook # 엑셀
from pytz import timezone
# import urlparse
import telegram
import requests
import datetime
import time
import ssl
import json
import re
import pymysql
import pymysql.cursors
from typing import List
from bs4 import BeautifulSoup
#from urllib.parse import urlparse
import urllib.parse as urlparse
import urllib.request


from requests import get  # to make GET request

# 로직 설명
# 1. Main()-> 각 회사별 함수를 통해 반복 (추후 함수명 일괄 변경 예정)
#   - checkNewArticle -> parse -> downloadFile -> Send 
# 2. 연속키의 경우 현재 .key로 저장
#   - 추후 heroku db로 처리 예정(MySQL)
#   - DB연결이 안되는 경우, Key로 처리할수 있도록 예외처리 반영
# 3. 최초 조회되는 게시판 혹은 Key값이 없는 경우 메세지를 발송하지 않음.
# 4. 테스트와 운영을 구분하여 텔레그램 발송 채널 ID 구분 로직 추가
#   - 어떻게 구분지을지 생각해봐야함
# 5. 메시지 발송 방법 변경 (봇 to 사용자 -> 채널에 발송)

############공용 상수############
# 메시지 발송 ID
CHAT_ID = '-1001431056975' # 운영 채널(증권사 신규 레포트 게시물 알림방)
# CHAT_ID = '-1001474652718' # 테스트 채널
# CHAT_ID = '-1001436418974' # 네이버 실시간 속보 뉴스 채널
# CHAT_ID = '-1001150510299' # 네이버 많이본 뉴스 채널
# CHAT_ID = '-1001472616534' # 아이투자


# DATABASE
CLEARDB_DATABASE_URL = 'mysql://b0464b22432146:290edeca@us-cdbr-east-03.cleardb.com/heroku_31ee6b0421e7ff9?reconnect=true'

# 게시글 갱신 시간
REFRESH_TIME = 600

# 회사이름
FIRM_NAME = (
    "이베스트 투자증권",    # 0
    "흥국증권",             # 1
    "상상인증권",           # 2
    "하나금융투자",          # 3
    "한양증권",              # 4
    "삼성증권",              # 5
    "교보증권"              # 6
    # "유안타증권",           # 4
)

# 게시판 이름
BOARD_NAME = (
    [ "이슈브리프" , "기업분석", "산업분석", "투자전략", "Quant" ], # 0
    [ "투자전략", "산업/기업분석" ],                            # 1
    [ "산업리포트", "기업리포트" ],                             # 2
    [ "산업분석", "기업분석", "Daily" ],                       # 3
    [ "기업분석", "산업 및 이슈분석" ],                          # 4
    [ "국내기업분석", "국내산업분석", "해외기업분석" ],              # 5
    [ " " ]                                                 # 6 (교보는 게시판 내 게시판 분류 사용)
    # [ "투자전략", "Report & Note", "해외주식" ],               # 4 => 유안타 데이터 보류 
)

EBEST_BOARD_NAME  = ["이슈브리프" , "기업분석", "산업분석", "투자전략", "Quant"]
HEUNGKUK_BOARD_NAME = ["투자전략", "산업/기업분석"]
SANGSANGIN_BOARD_NAME = ["산업리포트", "기업리포트"]
HANA_BOARD_NAME = ["산업분석", "기업분석", "Daily"]
HANYANG_BOARD_NAME = ["기업분석", "산업분석"]
HMSEC_BOARD_NAME = ["투자전략", "Report & Note", "해외주식"]
SAMSUNG_BOARD_NAME  = ["국내기업분석", "국내산업분석", "해외기업분석"]
# pymysql 변수
conn    = ''
cursor  = ''

# 연속키URL
NXT_KEY = ''
# 텔레그램 채널 발송 여부
SEND_YN = ''
# 첫번째URL 
FIRST_ARTICLE_URL = ''

# LOOP 인덱스 변수
SEC_FIRM_ORDER = 0 # 증권사 순번
ARTICLE_BOARD_ORDER = 0 # 게시판 순번

# 이모지
EMOJI_FIRE = u'\U0001F525'
EMOJI_PICK = u'\U0001F449'

# 연속키용 상수
FIRST_ARTICLE_INDEX = 0

# 메세지 전송용 레포트 제목(말줄임표 사용 증권사)
LIST_ARTICLE_TITLE = ''






def NAVERNews_checkNewArticle():
    global ARTICLE_BOARD_ORDER
    global SEC_FIRM_ORDER

    SEC_FIRM_ORDER      = 998
    ARTICLE_BOARD_ORDER = 998

    requests.packages.urllib3.disable_warnings()

    # 네이버 실시간 속보
    TARGET_URL_0 = 'http://wise.thewm.co.kr/ASP/Screener/data/Screener_Termtabledata.asp?market=0&industry=G0&size=0&workDT=20210305&termCount=4&currentPage=1&orderKey=P1&orderDirect=D&jsonParam=%5B%7B%22Group%22%3A%22I%22%2C%22SEQ%22%3A%222%22%2C%22MIN_VAL%22%3A%226096%22%2C%22MAX_VAL%22%3A%22200000%22%2C%22Ogb%22%3A%223%22%7D%2C%7B%22Group%22%3A%22P%22%2C%22SEQ%22%3A%221%22%2C%22MIN_VAL%22%3A%2210.00%22%2C%22MAX_VAL%22%3A%22100.00%22%2C%22Ogb%22%3A%221%22%7D%2C%7B%22Group%22%3A%22V%22%2C%22SEQ%22%3A%223%22%2C%22MIN_VAL%22%3A%221.00%22%2C%22MAX_VAL%22%3A%2241.00%22%2C%22Ogb%22%3A%221%22%7D%2C%7B%22Group%22%3A%22S%22%2C%22SEQ%22%3A%221%22%2C%22MIN_VAL%22%3A%22-1635%22%2C%22MAX_VAL%22%3A%22100.00%22%2C%22Ogb%22%3A%223%22%7D%5D'
    
    # 네이버 많이 본 뉴스
    TARGET_URL_1 = 'https://m.stock.naver.com/api/json/news/newsListJson.nhn?category=ranknews'
    
    TARGET_URL_TUPLE = (TARGET_URL_0, TARGET_URL_1)

    # URL GET
    for ARTICLE_BOARD_ORDER, TARGET_URL in enumerate(TARGET_URL_TUPLE):
        NAVERNews_parse(ARTICLE_BOARD_ORDER, TARGET_URL)
        time.sleep(5)
 
# JSON API 타입
def NAVERNews_parse():
    global NXT_KEY

    # TARGET_URL = 'http://wise.thewm.co.kr/ASP/Screener/data/Screener_Termtabledata.asp?market=0&industry=G0&size=0&workDT=20210305&termCount=4&currentPage=1&orderKey=P1&orderDirect=D&jsonParam=%5B%7B%22Group%22%3A%22I%22%2C%22SEQ%22%3A%222%22%2C%22MIN_VAL%22%3A%226096%22%2C%22MAX_VAL%22%3A%22200000%22%2C%22Ogb%22%3A%223%22%7D%2C%7B%22Group%22%3A%22P%22%2C%22SEQ%22%3A%221%22%2C%22MIN_VAL%22%3A%2210.00%22%2C%22MAX_VAL%22%3A%22100.00%22%2C%22Ogb%22%3A%221%22%7D%2C%7B%22Group%22%3A%22V%22%2C%22SEQ%22%3A%223%22%2C%22MIN_VAL%22%3A%221.00%22%2C%22MAX_VAL%22%3A%2241.00%22%2C%22Ogb%22%3A%221%22%7D%2C%7B%22Group%22%3A%22S%22%2C%22SEQ%22%3A%221%22%2C%22MIN_VAL%22%3A%22-1635%22%2C%22MAX_VAL%22%3A%22100.00%22%2C%22Ogb%22%3A%223%22%7D%5D'
    # TARGET_URL = 'http://wise.thewm.co.kr/ASP/Screener/data/Screener_Termtabledata.asp?market=0&industry=G0&size=0&workDT=20210305&termCount=3&currentPage=1&orderKey=P1&orderDirect=D&jsonParam=%5B%7B%22Group%22%3A%22I%22%2C%22SEQ%22%3A%222%22%2C%22MIN_VAL%22%3A%22100000.00%22%2C%22MAX_VAL%22%3A%22500000.00%22%2C%22Ogb%22%3A%221%22%7D%2C%7B%22Group%22%3A%22V%22%2C%22SEQ%22%3A%221%22%2C%22MIN_VAL%22%3A%221.00%22%2C%22MAX_VAL%22%3A%2230.00%22%2C%22Ogb%22%3A%221%22%7D%2C%7B%22Group%22%3A%22P%22%2C%22SEQ%22%3A%221%22%2C%22MIN_VAL%22%3A%225.00%22%2C%22MAX_VAL%22%3A%2250.00%22%2C%22Ogb%22%3A%221%22%7D%5D'
    TARGET_URL = 'http://wise.thewm.co.kr/ASP/Screener/data/Screener_Termtabledata.asp?market=0&industry=G0&size=0&workDT=20210311&termCount=2&currentPage=1&orderKey=V3&orderDirect=D&jsonParam=%5B%7B%22Group%22%3A%22I%22%2C%22SEQ%22%3A%222%22%2C%22MIN_VAL%22%3A%22100000.00%22%2C%22MAX_VAL%22%3A%22500000.00%22%2C%22Ogb%22%3A%221%22%7D%2C%7B%22Group%22%3A%22V%22%2C%22SEQ%22%3A%223%22%2C%22MIN_VAL%22%3A%221.00%22%2C%22MAX_VAL%22%3A%2210.00%22%2C%22Ogb%22%3A%221%22%7D%5D'

    request = urllib.request.Request(TARGET_URL)
    #검색 요청 및 처리
    response = urllib.request.urlopen(request)
    rescode = response.getcode()
    if rescode != 200 :return print("네이버 뉴스 접속이 원활하지 않습니다 ")

    CMP_PAGE_CNT = 10
    jres = json.loads(response.read().decode('utf-8'))
    #print(jres)
    TOTAL_CMP_CNT = jres['sAllCnt']
    TOTAL_PAGE_CNT = math.ceil(TOTAL_CMP_CNT / CMP_PAGE_CNT) # 페이지 수 이므로 정수가 아닌 경우 +1
    
    print('한 페이지에 회사 수는', CMP_PAGE_CNT , "건 입니다.")
    print('조건에 부합하는 회사 수는 ',TOTAL_CMP_CNT, "건 입니다.")
    print(TOTAL_PAGE_CNT)
    print("VAL 값은 우측 상단의 값임")
    print("반복코드는 나중에")
    
    file = open('hello.txt', 'w')    # hello.txt 파일을 쓰기 모드(w)로 열기. 파일 객체 반환
    
    
    NAVER_URL= 'https://finance.naver.com/item/main.nhn?code='
    jres = jres['resultList']
    for idx in range(1, TOTAL_PAGE_CNT+1):
        paging = 'currentPage='
        paging += str(idx)
        
        # print(TARGET_URL)
        request = urllib.request.Request(TARGET_URL)
        #검색 요청 및 처리
        response = urllib.request.urlopen(request)
        rescode = response.getcode()
        if rescode != 200 :return print("네이버 뉴스 접속이 원활하지 않습니다 ")
        jres = json.loads(response.read().decode('utf-8'))
        jres = jres['resultList']
        # print(idx)
        
        for r in jres:
            write = ''
            write += NAVER_URL + r['CMP_CD'] + '\t' +'종목명:' + r['CMP_NM_KOR'] + '\n'
            write += fnguide_parse(r['CMP_CD']) + '\n'
            print(write)
            file.write(write)      # 파일에 문자열 저장
            
            
        
        TARGET_URL = TARGET_URL.replace('currentPage='+ str(idx), 'currentPage='+ str(idx+1)  )
        
    file.close()                     # 파일 객체 닫기
    return
    FIRST_ARTICLE_TITLE = jres['newsList'][0]['tit'].strip()
    print('FIRST_ARTICLE_TITLE:',FIRST_ARTICLE_TITLE)

    # 연속키 데이터베이스화 작업
    # 연속키 데이터 저장 여부 확인 구간
    dbResult = DB_SelNxtKey(SEC_FIRM_ORDER, ARTICLE_BOARD_ORDER)
    if dbResult: # 1
        # 연속키가 존재하는 경우
        print('데이터베이스에 연속키가 존재합니다. ','(네이버 뉴스)')

    else: # 0
        # 연속키가 존재하지 않는 경우 => 첫번째 게시물 연속키 정보 데이터 베이스 저장
        print('데이터베이스에 ', '(네이버 뉴스)')
        NXT_KEY = DB_InsNxtKey(SEC_FIRM_ORDER, ARTICLE_BOARD_ORDER, FIRST_ARTICLE_TITLE)

    # NaverNews 게시판에 따른 URL 지정
    if ARTICLE_BOARD_ORDER == 0:category = 'flashnews'
    else:                      category = 'ranknews'

    nNewArticleCnt = 0
    sendMessageText = ''
    # JSON To List
    for news in jres['newsList']:
        LIST_ARTICLE_URL = 'https://m.stock.naver.com/news/read.nhn?category='+ category + '&officeId=' + news['oid'] + '&articleId=' + news['aid']
        LIST_ARTICLE_TITLE = news['tit'].strip()

        if ( NXT_KEY != LIST_ARTICLE_TITLE or NXT_KEY == '' ) and SEND_YN == 'Y':
            nNewArticleCnt += 1 # 새로운 게시글 수
            if len(sendMessageText) < 3500:
                sendMessageText += GetSendMessageText(INDEX = nNewArticleCnt ,ARTICLE_BOARD_NAME = '',ARTICLE_TITLE = LIST_ARTICLE_TITLE, ARTICLE_URL = LIST_ARTICLE_URL)
                print(len(sendMessageText))
            else:
                print("발송 게시물이 남았지만 최대 길이로 인해 중간 발송처리합니다.")
                sendText(sendMessageText)
                nNewArticleCnt = 0

        elif SEND_YN == 'N':
            print('###점검중 확인요망###')
        else:
            if nNewArticleCnt == 0:
                print('새로운 게시물을 모두 발송하였습니다.')
            else:
                sendText(sendMessageText)

            DB_UpdNxtKey(SEC_FIRM_ORDER, ARTICLE_BOARD_ORDER, FIRST_ARTICLE_TITLE)
            return True

    DB_UpdNxtKey(SEC_FIRM_ORDER, ARTICLE_BOARD_ORDER, FIRST_ARTICLE_TITLE) # 뉴스의 경우 연속 데이터가 다음 페이지로 넘어갈 경우 처리
    return True

def NAVERNews_downloadFile(LIST_ARTICLE_URL, LIST_ATTACT_FILE_NAME):
    global ATTACH_FILE_NAME
    ATTACH_FILE_NAME = LIST_ATTACT_FILE_NAME #BeautifulSoup(webpage.content, "html.parser").select_one('#contents > div > div.bbs_a_view > dl.b_bottom > dd > em:nth-child(1)> a').text.strip()
    
    DownloadFile(URL = LIST_ARTICLE_URL, FILE_NAME = ATTACH_FILE_NAME)
    time.sleep(5) # 모바일 알림을 받기 위해 8초 텀을 둠(loop 호출시)
    return True



# 최초 send함수
# URL(프리뷰해제) 발송 + 해당 레포트 pdf 발송
def send(ARTICLE_BOARD_NAME , ARTICLE_TITLE , ARTICLE_URL): # 파일의 경우 전역변수로 처리 (downloadFile 함수)
    global CHAT_ID

    print('send()')
    DISABLE_WEB_PAGE_PREVIEW = True # 메시지 프리뷰 여부 기본값 설정

    # 실제 전송할 메시지 작성
    sendMessageText = ''
    sendMessageText += GetSendMessageTitle(ARTICLE_TITLE)
    sendMessageText += ARTICLE_TITLE + "\n"
    sendMessageText += EMOJI_PICK + ARTICLE_URL 

    #생성한 텔레그램 봇 정보 assign (@ebest_noti_bot)
    my_token_key = '1372612160:AAHVyndGDmb1N2yEgvlZ_DmUgShqk2F0d4w'
    bot = telegram.Bot(token = my_token_key)

    #생성한 텔레그램 봇 정보 출력
    #me = bot.getMe()
    #print('텔레그램 채널 정보 :',me)

    if SEC_FIRM_ORDER == 999 or SEC_FIRM_ORDER == 998 or SEC_FIRM_ORDER == 997 : # 매매동향의 경우 URL만 발송하여 프리뷰 처리 
        DISABLE_WEB_PAGE_PREVIEW = False


    # if SEC_FIRM_ORDER == 998:
    #     if  ARTICLE_BOARD_ORDER == 0 : 
    #         CHAT_ID = '-1001436418974' # 네이버 실시간 속보 뉴스 채널
    #     else:
    #         CHAT_ID = '-1001150510299' # 네이버 많이본 뉴스 채널
    # elif SEC_FIRM_ORDER == 997:
    #         CHAT_ID = '-1001472616534' # 아이투자
    # else:
    #     CHAT_ID = '-1001431056975' # 운영 채널(증권사 신규 레포트 게시물 알림방)

    bot.sendMessage(chat_id = GetSendChatId(), text = sendMessageText, disable_web_page_preview = DISABLE_WEB_PAGE_PREVIEW)

    if DISABLE_WEB_PAGE_PREVIEW: # 첨부파일이 있는 경우 => 프리뷰는 사용하지 않음
        try:
            time.sleep(1) # 메시지 전송 텀을 두어 푸시를 겹치지 않게 함
            bot.sendDocument(chat_id = GetSendChatId(), document = open(ATTACH_FILE_NAME, 'rb'))
            os.remove(ATTACH_FILE_NAME) # 파일 전송 후 PDF 삭제
        except:
            return
    
    time.sleep(8) # 모바일 알림을 받기 위해 8초 텀을 둠(loop 호출시)

# URL 발신용 전용 함수 : ex) 네이버 뉴스
def sendURL(ARTICLE_BOARD_NAME , ARTICLE_TITLE , ARTICLE_URL): # 파일의 경우 전역변수로 처리 (downloadFile 함수)
    global CHAT_ID

    print('sendURL()')

    # 실제 전송할 메시지 작성
    sendMessageText = ''
    sendMessageText += GetSendMessageTitle(ARTICLE_TITLE)
    sendMessageText += ARTICLE_TITLE + "\n"
    sendMessageText += EMOJI_PICK + ARTICLE_URL 

    #생성한 텔레그램 봇 정보 assign (@ebest_noti_bot)
    my_token_key = '1372612160:AAHVyndGDmb1N2yEgvlZ_DmUgShqk2F0d4w'
    bot = telegram.Bot(token = my_token_key)

    #생성한 텔레그램 봇 정보 출력
    #me = bot.getMe()
    #print('텔레그램 채널 정보 :',me)

    # if SEC_FIRM_ORDER == 998:
    #     if  ARTICLE_BOARD_ORDER == 0 : 
    #         CHAT_ID = '-1001436418974' # 네이버 실시간 속보 뉴스 채널
    #     else:
    #         CHAT_ID = '-1001150510299' # 네이버 많이본 뉴스 채널
    # elif SEC_FIRM_ORDER == 997:
    #         CHAT_ID = '-1001472616534' # 아이투자
    # else:
    #     CHAT_ID = '-1001431056975' # 운영 채널(증권사 신규 레포트 게시물 알림방)

    bot.sendMessage(chat_id = GetSendChatId(), text = sendMessageText)
    
    time.sleep(8) # 모바일 알림을 받기 위해 8초 텀을 둠(loop 호출시)

def sendPhoto(ARTICLE_URL): # 파일의 경우 전역변수로 처리 (downloadFile 함수)
    print('sendPhoto()')

    #생성한 텔레그램 봇 정보 assign (@ebest_noti_bot)
    my_token_key = '1372612160:AAHVyndGDmb1N2yEgvlZ_DmUgShqk2F0d4w'
    bot = telegram.Bot(token = my_token_key)

    bot.sendPhoto(chat_id = GetSendChatId(), photo = ARTICLE_URL)
    time.sleep(8) # 모바일 알림을 받기 위해 8초 텀을 둠(loop 호출시)
    return True

def sendText(sendMessageText): # 가공없이 텍스트를 발송합니다.
    global CHAT_ID

    print('sendText()')

    #생성한 텔레그램 봇 정보 assign (@ebest_noti_bot)
    my_token_key = '1372612160:AAHVyndGDmb1N2yEgvlZ_DmUgShqk2F0d4w'
    bot = telegram.Bot(token = my_token_key)

    # if SEC_FIRM_ORDER == 998:
    #     if  ARTICLE_BOARD_ORDER == 0 : 
    #         CHAT_ID = '-1001436418974' # 네이버 실시간 속보 뉴스 채널
    #     else:
    #         CHAT_ID = '-1001150510299' # 네이버 많이본 뉴스 채널
    # elif SEC_FIRM_ORDER == 997:
    #         CHAT_ID = '-1001472616534' # 아이투자
    # else:
    #     CHAT_ID = '-1001431056975' # 운영 채널(증권사 신규 레포트 게시물 알림방)

    bot.sendMessage(chat_id = GetSendChatId(), text = sendMessageText, disable_web_page_preview = True, parse_mode = "Markdown")
    
    time.sleep(8) # 모바일 알림을 받기 위해 8초 텀을 둠(loop 호출시)

# URL에 파일명을 사용할때 한글이 포함된 경우 인코딩처리 로직 추가 
def DownloadFile(URL, FILE_NAME):
    global ATTACH_FILE_NAME
    print("DownloadFile()")

    if SEC_FIRM_ORDER == 6: # 교보증권 예외 로직
        # 로직 사유 : 레포트 첨부파일명에 한글이 포함된 경우 URL처리가 되어 있지 않음
        CONVERT_URL = URL 
        for c in URL: # URL내 한글이 있는 경우 인코딩 처리(URL에 파일명을 이용하여 조합함)
            # 코드셋 기준 파이썬:UTF-8 . 교보증권:EUC-KR
            # 1. 주소에서 한글 문자를 판별
            # 2. 해당 문자를 EUC-KR로 변환후 URL 인코딩
            print("##",c , "##", ord('가') <= ord(c) <= ord('힣') )
            if ord('가') <= ord(c) <= ord('힣'): 
                c_encode = c.encode('euc-kr')
                CONVERT_URL = CONVERT_URL.replace(c, urlparse.quote(c_encode) )
                print(CONVERT_URL)

        if URL != CONVERT_URL: 
            print("기존 URL에 한글이 포함되어 있어 인코딩처리함")
            print("CONVERT_URL", CONVERT_URL)
            URL = CONVERT_URL

    ATTACH_FILE_NAME = re.sub('[\/:*?"<>|]','',FILE_NAME) # 저장할 파일명 : 파일명으로 사용할수 없는 문자 삭제 변환
    print('convert URL:',URL)
    print('convert ATTACH_FILE_NAME:',ATTACH_FILE_NAME)
    with open(ATTACH_FILE_NAME, "wb")as file:  # open in binary mode
        response = get(URL, verify=False)     # get request
        file.write(response.content) # write to file
        
    return True

def GetSendMessageText(INDEX, ARTICLE_BOARD_NAME , ARTICLE_TITLE , ARTICLE_URL):

    # 실제 전송할 메시지 작성
    sendMessageText = ''
    # 발신 게시판 종류
    if INDEX == 1:
        sendMessageText += GetSendMessageTitle(ARTICLE_TITLE) + "\n"
    # 게시글 제목(굵게)
    sendMessageText += "**" + ARTICLE_TITLE + "**" + "\n"
    # 원문 링크
    sendMessageText += EMOJI_PICK  + "[원문링크(클릭)]" + "("+ ARTICLE_URL + ")"
    sendMessageText += "\n" + "\n"

    return sendMessageText

def GetSendMessageTitle(ARTICLE_TITLE):

    SendMessageTitle = ''
    if SEC_FIRM_ORDER == 999:
        msgFirmName = "매매동향"
        ARTICLE_BOARD_NAME = ''
        if  "최종치" in ARTICLE_TITLE:
            print('sedaily의 매매동향 최종치 집계 데이터는 메시지 발송을 하지 않습니다.') # 장마감 최종치는 발송 안함
            return 
    elif SEC_FIRM_ORDER == 998:
        msgFirmName = "네이버 - "
        if  ARTICLE_BOARD_ORDER == 0 :
            ARTICLE_BOARD_NAME = "실시간 뉴스 속보"
        else:
            ARTICLE_BOARD_NAME = "가장 많이 본 뉴스"
    elif SEC_FIRM_ORDER == 997:
        msgFirmName = "아이투자 - "
    else:
        msgFirmName = FIRM_NAME[SEC_FIRM_ORDER] + " - "
        if SEC_FIRM_ORDER != 6: 
            ARTICLE_BOARD_NAME = BOARD_NAME[SEC_FIRM_ORDER][ARTICLE_BOARD_ORDER]

    SendMessageTitle += EMOJI_FIRE + msgFirmName + ARTICLE_BOARD_NAME + EMOJI_FIRE + "\n"
    
    return SendMessageTitle


def GetSendChatId():
    SendMessageChatId = 0
    if SEC_FIRM_ORDER == 998:
        if  ARTICLE_BOARD_ORDER == 0 : 
            SendMessageChatId = '-1001436418974' # 네이버 실시간 속보 뉴스 채널
        else:
            SendMessageChatId = '-1001150510299' # 네이버 많이본 뉴스 채널
    elif SEC_FIRM_ORDER == 997:
            SendMessageChatId = '-1001472616534' # 아이투자
    else:
        SendMessageChatId = '-1001431056975' # 운영 채널(증권사 신규 레포트 게시물 알림방)
    
    return SendMessageChatId

def MySQL_Open_Connect():
    global conn
    global cursor
    
    # clearDB 
    # url = urlparse.urlparse(os.environ['CLEARDB_DATABASE_URL'])
    url = urlparse.urlparse('mysql://b0464b22432146:290edeca@us-cdbr-east-03.cleardb.com/heroku_31ee6b0421e7ff9?reconnect=true')
    conn = pymysql.connect(host=url.hostname, user=url.username, password=url.password, charset='utf8', db=url.path.replace('/', ''), cursorclass=pymysql.cursors.DictCursor, autocommit=True)
    cursor = conn.cursor()
    return cursor

def DB_SelNxtKey(SEC_FIRM_ORDER, ARTICLE_BOARD_ORDER):
    global NXT_KEY
    global SEND_YN
    global conn
    global cursor

    cursor = MySQL_Open_Connect()
    dbQuery = "SELECT * FROM NXT_KEY WHERE 1=1 AND  SEC_FIRM_ORDER = %s   AND ARTICLE_BOARD_ORDER = %s "
    dbResult = cursor.execute(dbQuery, (SEC_FIRM_ORDER, ARTICLE_BOARD_ORDER))
    rows = cursor.fetchall()
    for row in rows:
        print('####DB조회된 연속키####', end='\n')
        print('SEC_FIRM_ORDER',row['SEC_FIRM_ORDER'], 'ARTICLE_BOARD_ORDER',row['ARTICLE_BOARD_ORDER'], 'NXT_KEY',row['NXT_KEY'])
        NXT_KEY = row['NXT_KEY']
        SEND_YN = row['SEND_YN']
    conn.close()
    return dbResult

def DB_InsNxtKey(SEC_FIRM_ORDER, ARTICLE_BOARD_ORDER, FIRST_NXT_KEY):
    global NXT_KEY
    global conn
    global cursor
    cursor = MySQL_Open_Connect()
    dbQuery = "INSERT INTO NXT_KEY (SEC_FIRM_ORDER, ARTICLE_BOARD_ORDER, NXT_KEY)VALUES ( %s, %s, %s);"
    cursor.execute(dbQuery, ( SEC_FIRM_ORDER, ARTICLE_BOARD_ORDER, FIRST_NXT_KEY ))
    NXT_KEY = FIRST_NXT_KEY
    conn.close()
    return NXT_KEY

def DB_UpdNxtKey(SEC_FIRM_ORDER, ARTICLE_BOARD_ORDER, FIRST_NXT_KEY):
    global NXT_KEY
    cursor = MySQL_Open_Connect()
    dbQuery = "UPDATE NXT_KEY SET NXT_KEY = %s WHERE 1=1 AND  SEC_FIRM_ORDER = %s   AND ARTICLE_BOARD_ORDER = %s;"
    dbResult = cursor.execute(dbQuery, ( FIRST_NXT_KEY, SEC_FIRM_ORDER, ARTICLE_BOARD_ORDER ))
    if dbResult:
        NXT_KEY = FIRST_NXT_KEY
    conn.close()
    return dbResult


def fnguide_parse(*args):
    global NXT_KEY
    global LIST_ARTICLE_TITLE

    pattern = ''
    CODE = ''
    for pattern in args:
        if len(pattern) > 0 :
            CODE =  pattern

    TARGET_URL = 'http://comp.fnguide.com/SVO2/ASP/SVD_Main.asp?MenuYn=Y&gicode=A'
    TARGET_URL += CODE
    # 005930
    webpage = requests.get(TARGET_URL, verify=False)

    # HTML parse
    soup = BeautifulSoup(webpage.content, "html.parser")
    data_cmp_nm = soup.select_one('#giName').text
    data_cmp_code = soup.select_one('#compBody > div.section.ul_corpinfo > div.corp_group1 > h2').text
    data_stxt1 = soup.select_one('#compBody > div.section.ul_corpinfo > div.corp_group1 > p > span.stxt.stxt1').text
    data_stxt2 = soup.select_one('#strMarketTxt').text
    data_Per = soup.select_one('#corp_group2 > dl:nth-child(1) > dd').text
    data_fwdPer = soup.select_one('#corp_group2 > dl:nth-child(2) > dd').text
    data_dividendYield = soup.select_one('#corp_group2 > dl:nth-child(5) > dd').text
    data_cmp_info = soup.select_one('#bizSummaryContent').text
    #data_ROE = soup.select_one('#svdMainGrid10D > table > tbody > tr:nth-child(7) > td:nth-child(2)')#.text
    r = ''
    r += TARGET_URL + '\n'
    r += '==============================================================' + '\n'
    r += '종목명: ' + data_cmp_nm                                + '\n'
    r += '종목코드: ' + data_cmp_code                            + '\n'
    r += '업종: ' + data_stxt1                                      + '\t' + data_stxt2 + '\n'
    r += 'per(FY0): ' + data_Per                                      + '\n'
    r += '12m fwd per: ' + data_fwdPer                           + '\n'
    r += '시가배당 수익률 '  + data_dividendYield                + '\n'
    r += '기업개요:' + data_cmp_info                 + '\n'
    r += '==============================================================' + '\n'
    #print('ROE', data_ROE)
    
    
    return r

def excel_write_title(*args):
    global write_wb
    global write_ws
    
    pattern = ''
    CODE = ''
    for pattern in args:
        if len(pattern) ==  0 or pattern ==  NoneType :
                # 엑셀파일 쓰기
                write_wb = Workbook()
                # Sheet1에다 입력
                write_ws = write_wb.active
                # 타이틀
                write_ws.cell(1, 1, '링크')
                write_ws.cell(1, 2, '종목명')
                write_ws.cell(1, 3, '종목코드')
                write_ws.cell(1, 4, 'PER')
                write_ws.cell(1, 5, 'fwd-PER')
                write_ws.cell(1, 6, '시가배당')




    # 엑셀파일 쓰기
    # write_wb = Workbook()

    # 이름이 있는 시트를 생성
    # write_ws = write_wb.create_sheet('생성시트')

    # # Sheet1에다 입력
    # write_ws = write_wb.active
    # # 타이틀
    # write_ws['A1'] = '숫자'
    # write_ws['B1'] = '종목명'

    #행 단위로 추가
    write_ws.append([1,2,3])

    #셀 단위로 추가
    write_ws.cell(5, 5, '5행5열')
    write_wb.save("숫자.xlsx")

    # 출처 https://myjamong.tistory.com/51

def main():
    global SEC_FIRM_ORDER  # 증권사 순번
    print('########Program Start Run########')

    # SEC_FIRM_ORDER는 임시코드 추후 로직 추가 예정 
    while True:
        
        # fnguide_parse('005930')
        # excel_write()
        NAVERNews_parse()
        # return
        # print("EBEST_checkNewArticle()=> 새 게시글 정보 확인") # 0
        # EBEST_checkNewArticle()
        
        # print("HeungKuk_checkNewArticle()=> 새 게시글 정보 확인") # 1
        # HeungKuk_checkNewArticle()

        # print("SangSangIn_checkNewArticle()=> 새 게시글 정보 확인") # 2
        # SangSangIn_checkNewArticle()

        # print("HANA_checkNewArticle()=> 새 게시글 정보 확인") # 3
        # HANA_checkNewArticle()

        # print("HANYANG_checkNewArticle()=> 새 게시글 정보 확인") # 4
        # HANYANG_checkNewArticle()

        # print("Samsung_checkNewArticle()=> 새 게시글 정보 확인") # 5
        # Samsung_checkNewArticle()

        # print("KyoBo_checkNewArticle()=> 새 게시글 정보 확인") # 6
        # KyoBo_checkNewArticle()

        # print("Itooza_checkNewArticle()=> 새 게시글 정보 확인") # 997 미활성
        # Itooza_checkNewArticle()

        # print("NAVERNews_checkNewArticle()=> 새 게시글 정보 확인") # 998 미활성
        # NAVERNews_checkNewArticle()

        # print("SEDAILY_checkNewArticle()=> 새 게시글 정보 확인") # 999
        # SEDAILY_checkNewArticle()

        # print("YUANTA_checkNewArticle()=> 새 게시글 정보 확인") # 4 가능여부 불확실 => 보류
        # YUANTA_checkNewArticle()
        print('######',REFRESH_TIME,'초 후 게시글을 재 확인 합니다.######')
        time.sleep(REFRESH_TIME)

if __name__ == "__main__":
	main()
