import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
from datetime import datetime, timedelta
from urllib.parse import urlencode
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm
import time
import random
import re
from calendar import monthrange

# 유효하지 않은 문자 제거 함수
def clean_string(value):
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    return ILLEGAL_CHARACTERS_RE.sub('', value)

# URL 설정 함수 (특정 날짜 범위로 설정)
def get_url(start_date, end_date):
    params = {
        'fr_dt': start_date.strftime('%Y%m%d'),
        'to_dt': end_date.strftime('%Y%m%d'),
        'stext': '',
        'check': 'all',
        'sortOrd': '5',
        'sortAD': 'A',
        '_': '1718850387871'
    }
    base_url = "https://comp.fnguide.com/SVO2/ASP/SVD_Report_Summary_Data.asp"
    url = f"{base_url}?{urlencode(params)}"
    return url

# 한글 포함 여부 확인 함수
def contains_korean(text):
    return bool(re.search(r'[\u3131-\u3163\uac00-\ud7a3]', text))

# 데이터 크롤링 함수
def fetch_data(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')

    data = []
    rows = soup.find_all('tr', {'class': ['zigbg_in', '']})
    for row in rows:
        columns = row.find_all('td')
        if columns:
            try:
                date = clean_string(columns[0].text.strip().replace('\n', '')) if columns[0] else ''
                code_name_element = columns[1].find('a')
                code_name = clean_string(code_name_element.text.strip()) if code_name_element else ''
                code_element = columns[1].find('span', class_='txt1')
                code = clean_string(code_element.text.strip()) if code_element else ''
                company_name = code_name.replace(f" {code}", "") if code else code_name
                title_element = columns[1].find('span', class_='txt2')
                title = clean_string(title_element.text.strip().lstrip('-')) if title_element else ''
                
                # 한글이 포함되지 않은 레포트 제목은 제외
                if not contains_korean(title):
                    continue

                summary_parts = columns[1].find_all('dd')
                summary = clean_string(" ".join([part.text.strip() for part in summary_parts])) if summary_parts else ''
                opinion = clean_string(columns[2].text.strip()) if columns[2] else ''
                target_price = clean_string(columns[3].text.strip()) if columns[3] else ''
                current_price = clean_string(columns[4].text.strip()) if columns[4] else ''
                analyst_info = columns[5].find('span', class_='gpbox')
                brokerage = ''
                analyst = ''
                if analyst_info and hasattr(analyst_info, 'contents') and len(analyst_info.contents) > 1:
                    brokerage = clean_string(analyst_info.contents[0].strip()) if isinstance(analyst_info.contents[0], str) else ''
                    analyst = clean_string(analyst_info.contents[1].strip()) if isinstance(analyst_info.contents[1], str) else ''

                data.append([date, code, company_name, title, summary, opinion, target_price, current_price, analyst, brokerage])
            except Exception as e:
                print(f"Error processing row: {row}")
                print(f"Exception: {e}")

    df = pd.DataFrame(data, columns=["발행일자", "종목코드", "종목명", "레포트제목", "Summary", "투자의견", "목표주가", "전일종가", "애널리스트", "증권사"])
    return df

# 중복 데이터 체크 및 제거 함수
def remove_duplicates(new_df, existing_df):
    combined_df = pd.concat([existing_df, new_df])
    combined_df.drop_duplicates(inplace=True, ignore_index=True)
    return combined_df

# 엑셀 파일에 데이터 누적 저장 함수
def save_to_excel(df, filename, test_mode=False, separate_sheets=True):
    if test_mode and os.path.exists(filename):
        os.remove(filename)
    
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if separate_sheets:
        # 엑셀 파일에 월별 시트로 저장
        for date, group in df.groupby(df['발행일자']):
            sanitized_month = date.split('/')
            sanitized_month = f"{sanitized_month[0]}-{sanitized_month[1]}"
            if sanitized_month not in wb.sheetnames:
                ws = wb.create_sheet(title=sanitized_month)
                if test_mode:
                    for r in dataframe_to_rows(group.head(2), index=False, header=True):  # 테스트 모드에서 2개만 저장
                        ws.append(r)
                else:
                    for r in dataframe_to_rows(group, index=False, header=True):
                        ws.append(r)
            else:
                ws = wb[sanitized_month]
                if test_mode:
                    for r in dataframe_to_rows(group.head(2), index=False, header=False):  # 테스트 모드에서 2개만 저장
                        ws.append(r)
                else:
                    for r in dataframe_to_rows(group, index=False, header=False):
                        ws.append(r)
    else:
        # 모든 데이터를 한 시트에 저장
        year = df['발행일자'].iloc[0][:4]
        sheet_name = year
        ws = wb.create_sheet(title=sheet_name)

        # 월별로 데이터를 2건씩 추가
        for month in range(1, 13):
            month_str = f"{year}/{month:02d}"
            month_data = df[df['발행일자'].str.startswith(month_str)]
            if test_mode:
                month_data = month_data.head(2)  # 테스트 모드에서 2개만 저장
            if month_data.empty:
                continue
            for r in dataframe_to_rows(month_data, index=False, header=False):
                ws.append(r)

    wb.save(filename)

# 과거 데이터 수집 함수
def collect_historical_data(year, separate_sheets, test_mode=False):
    filename = f'financial_data_{year}.xlsx'
    if test_mode:
        filename = f'test_financial_data_{year}.xlsx'

    existing_months = set()
    if os.path.exists(filename):
        wb = load_workbook(filename)
        existing_months = set(wb.sheetnames)

    with tqdm(total=12, desc=f'Collecting data for {year}') as pbar:
        for month in range(1, 13):
            start_date = datetime(year, month, 1)
            _, last_day = monthrange(year, month)
            end_date = datetime(year, month, last_day)

            sanitized_month = f"{year}-{month:02d}"
            if sanitized_month in existing_months:
                pbar.update(1)
                continue

            url = get_url(start_date, end_date)
            df = fetch_data(url)

            if os.path.exists(filename):
                existing_df = pd.read_excel(filename, sheet_name=None)
                for sheet_name, sheet_df in existing_df.items():
                    if sheet_df.shape[1] == 9:
                        sheet_df.columns = ["발행일자", "종목코드", "종목명", "레포트제목", "Summary", "투자의견", "목표주가", "전일종가", "애널리스트"]
                    elif sheet_df.shape[1] == 10:
                        sheet_df.columns = ["발행일자", "종목코드", "종목명", "레포트제목", "Summary", "투자의견", "목표주가", "전일종가", "애널리스트", "증권사"]
                    df = remove_duplicates(df, sheet_df)

            print(df)  # 현재 수집된 데이터를 출력
            save_to_excel(df, filename, test_mode=test_mode, separate_sheets=separate_sheets)

            pbar.update(1)
            
            # 0.1~0.5초 사이의 무작위 딜레이 추가
            time.sleep(random.uniform(0.1, 0.5))

# 메인 함수
def main():
    year = int(input("출력연도를 입력해주세요: "))

    separate_sheets_input = input("월별로 시트를 분리할까요? (yes/no): ").strip().lower()
    separate_sheets = 'y' in separate_sheets_input

    test_mode_input = input("Is this a test run? (yes/no): ").strip().lower()
    test_mode = 'y' in test_mode_input

    collect_historical_data(year, separate_sheets, test_mode)

if __name__ == "__main__":
    main()
