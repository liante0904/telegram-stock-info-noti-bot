import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
from datetime import datetime, timedelta
from urllib.parse import urlencode
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from tqdm import tqdm
import time
import random
import re

# 유효하지 않은 문자 제거 함수
def clean_string(value):
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    return ILLEGAL_CHARACTERS_RE.sub('', value)

# URL 설정 함수 (특정 날짜 범위로 설정)
def get_url(start_date, end_date):
    params = {
        'fr_dt': start_date.strftime('%Y%m%d'),
        'to_dt': end_date.strftime('%Y%m%d'),
        'stext': '',
        'check': 'all',
        'sortOrd': '5',
        'sortAD': 'A',
        '_': '1718850387871'
    }
    base_url = "https://comp.fnguide.com/SVO2/ASP/SVD_Report_Summary_Data.asp"
    url = f"{base_url}?{urlencode(params)}"
    return url

# 데이터 크롤링 함수
def fetch_data(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')

    data = []
    rows = soup.find_all('tr', {'class': ['zigbg_in', '']})
    for row in rows:
        columns = row.find_all('td')
        if columns:
            try:
                date = clean_string(columns[0].text.strip().replace('\n', '')) if columns[0] else ''
                code_name_element = columns[1].find('a')
                code_name = clean_string(code_name_element.text.strip()) if code_name_element else ''
                code_element = columns[1].find('span', class_='txt1')
                code = clean_string(code_element.text.strip()) if code_element else ''
                company_name = code_name.replace(f" {code}", "") if code else code_name
                title_element = columns[1].find('span', class_='txt2')
                title = clean_string(title_element.text.strip().lstrip('-')) if title_element else ''
                summary_parts = columns[1].find_all('dd')
                summary = clean_string(" ".join([part.text.strip() for part in summary_parts])) if summary_parts else ''
                opinion = clean_string(columns[2].text.strip()) if columns[2] else ''
                target_price = clean_string(columns[3].text.strip()) if columns[3] else ''
                current_price = clean_string(columns[4].text.strip()) if columns[4] else ''
                analyst_info = columns[5].find('span', class_='gpbox')
                brokerage = ''
                analyst = ''
                if analyst_info and hasattr(analyst_info, 'contents') and len(analyst_info.contents) > 1:
                    brokerage = clean_string(analyst_info.contents[0].strip()) if isinstance(analyst_info.contents[0], str) else ''
                    analyst = clean_string(analyst_info.contents[1].strip()) if isinstance(analyst_info.contents[1], str) else ''

                data.append([date, code, company_name, title, summary, opinion, target_price, current_price, analyst, brokerage])
            except Exception as e:
                print(f"Error processing row: {row}")
                print(f"Exception: {e}")

    df = pd.DataFrame(data, columns=["발행일자", "종목코드", "종목명", "레포트제목", "Summary", "투자의견", "목표주가", "전일종가", "애널리스트", "증권사"])
    return df

# 중복 데이터 체크 및 제거 함수
def remove_duplicates(new_df, existing_df):
    combined_df = pd.concat([existing_df, new_df])
    combined_df.drop_duplicates(inplace=True, ignore_index=True)
    return combined_df

# 엑셀 파일에 데이터 누적 저장 및 필터 추가 함수
def save_to_excel(df, filename):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        if '통합' in wb.sheetnames:
            wb.remove(wb['통합'])
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # 엑셀 파일에 월별 시트로 저장
    for date, group in df.groupby(df['발행일자'].str[:6]):
        month = date[:6]
        sanitized_month = month.replace('/', '-')
        if sanitized_month not in wb.sheetnames:
            ws = wb.create_sheet(title=sanitized_month)
            for r in dataframe_to_rows(group, index=False, header=True):
                ws.append(r)
        else:
            ws = wb[sanitized_month]
            for r in dataframe_to_rows(group, index=False, header=False):
                ws.append(r)

    # 통합 시트에 데이터 추가
    combined_df = pd.concat([pd.DataFrame(ws.values) for ws in wb.worksheets if ws.title != '통합'])
    combined_df.columns = ["발행일자", "종목코드", "종목명", "레포트제목", "Summary", "투자의견", "목표주가", "전일종가", "애널리스트", "증권사"]

    combined_ws = wb.create_sheet(title='통합', index=0)
    for r in dataframe_to_rows(combined_df, index=False, header=True):
        combined_ws.append(r)

    wb.save(filename)

# 과거 데이터 수집 함수
def collect_historical_data(start_date, end_date):
    filename = f'financial_data_{start_date.year}.xlsx'

    existing_months = set()
    if os.path.exists(filename):
        wb = load_workbook(filename)
        existing_months = set(wb.sheetnames) - {'통합'}

    with tqdm(total=12, desc=f'Collecting data for {start_date.year}') as pbar:
        while start_date <= end_date:
            month = start_date.strftime('%Y%m')
            sanitized_month = month.replace('/', '-')
            if sanitized_month in existing_months:
                start_date = (start_date + timedelta(days=30)).replace(day=1)
                pbar.update(1)
                continue

            month_end = start_date + timedelta(days=30)
            if month_end > end_date:
                month_end = end_date

            url = get_url(start_date, month_end)
            df = fetch_data(url)

            if os.path.exists(filename):
                existing_df = pd.read_excel(filename, sheet_name=None)
                for sheet_name, sheet_df in existing_df.items():
                    if sheet_name != '통합':
                        sheet_df.columns = ["발행일자", "종목코드", "종목명", "레포트제목", "Summary", "투자의견", "목표주가", "전일종가", "애널리스트", "증권사"]
                        df = remove_duplicates(df, sheet_df)

            print(df)  # 현재 수집된 데이터를 출력
            save_to_excel(df, filename)

            start_date = month_end + timedelta(days=1)
            pbar.update(1)
            
            # 0.1~0.5초 사이의 무작위 딜레이 추가
            time.sleep(random.uniform(0.1, 0.5))

# 메인 함수
def main():
    today = datetime.today()
    start_date = today.replace(day=1)
    end_date = today

    collect_historical_data(start_date, end_date)

if __name__ == "__main__":
    main()
